#!/usr/bin/env python3

import sys, warnings
import openpyxl
from .gridformatter import GridFormatter, GridFormatterWithHeader

def print_data(obj):
    for attr in dir(obj):
        if not attr.startswith("_"):
            try:
                print(attr, "=", getattr(obj, attr))
            except NotImplementedError:
                pass
            print()
            

class WorkbookWriter:
    def __init__(self, fn):
        with open(fn, "rb") as f:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self.workbook = openpyxl.load_workbook(f)
        self.style_cells = {}

    def get_cell_text(self, cell):
        text = str(cell.value) if cell.value is not None else ""
        if cell.has_style:
            text += "*" * cell.style_id
            if cell.style_id not in self.style_cells:
                self.style_cells[cell.style_id] = cell
        return text

    def get_color_description(self, color):
        rgb = color.rgb
        if isinstance(rgb, str) and rgb != "00000000":
            return rgb

    def get_fill_description(self, fill):
        color_desc = self.get_color_description(fill.start_color)
        desc = "fill " + str(fill.fill_type)
        if color_desc:
            desc += " (" + color_desc + ")"
        return desc
        
    def get_font_description(self, font):
        desc = font.name + " " + str(int(font.size))
        parts = []
        if font.b:
            parts.append("bold")
        if font.i:
            parts.append("italic")
        if font.u:
            parts.append("underline")
        color_desc = self.get_color_description(font.color)
        if color_desc:
            parts.append(color_desc)
        if parts:
            desc += " (" + ",".join(parts) + ")"
        return desc
        
    def get_border_description(self, border):
        attrs = [ "top", "bottom", "left", "right", "diagonal"]
        parts = []
        for attr in attrs:
            data = getattr(border, attr)
            if data.style:
                desc = attr + " " + data.style
                color_desc = self.get_color_description(data.color)
                if color_desc:
                    desc += color_desc
                parts.append(desc)
        if len(parts) > 0:
            return "border " + " + ".join(parts)
        
    def get_style_description(self, cell):
        parts = [ self.get_font_description(cell.font) ]
        if cell.fill.fill_type:
            parts.append(self.get_fill_description(cell.fill))
        border_desc = self.get_border_description(cell.border)
        if border_desc:
            parts.append(border_desc)
        return ", ".join(parts)

    def get_sheet_description(self, sheet):
        sheet_desc = "Sheet '" + sheet.title + "' - " + str(sheet.max_row) + " rows " + str(sheet.max_column) + " columns"
        if sheet.sheet_properties.tabColor:
            color_desc = self.get_color_description(sheet.sheet_properties.tabColor)
            if color_desc:
                sheet_desc += ", tab color " + color_desc
        return sheet_desc

    def write(self):
        for sheet in self.workbook.worksheets:
            print(self.get_sheet_description(sheet))
            header_rows, body_rows = [], []
            in_body = False
            prev_data_types = None
            for row in sheet.iter_rows():
                datarow = [ self.get_cell_text(cell) for cell in row ]
                curr_data_types = [ type(cell.value) for cell in row ]
                if not in_body and prev_data_types and curr_data_types != prev_data_types:
                    in_body = True
                if in_body:
                    body_rows.append(datarow)
                else:
                    header_rows.append(datarow)
                prev_data_types = curr_data_types
            if body_rows:
                formatter = GridFormatterWithHeader(header_rows, body_rows, sheet.max_column)
            else:
                formatter = GridFormatter(header_rows, sheet.max_column)
            print(formatter)
            if not body_rows:
                print()
                
        for style_id, cell in self.style_cells.items():
            print("*" * style_id, self.get_style_description(cell))
            print()
   
def main_cli():
    last_file = max(sorted(sys.argv[1:]))
    writer = WorkbookWriter(last_file)
    writer.write()
                
if __name__ == '__main__':
    main_cli()