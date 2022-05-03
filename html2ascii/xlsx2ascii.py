#!/usr/bin/env python3

import sys
import openpyxl
from gridformatter import GridFormatter, GridFormatterWithHeader

if __name__ == '__main__':
    last_file = max(sorted(sys.argv[1:]))
    with open(last_file, "rb") as f:
        wb_obj = openpyxl.load_workbook(f)
        for sheet in wb_obj.worksheets:
            print("Sheet", "'" + sheet.title + "' -", sheet.max_row, "rows", sheet.max_column, "columns")
            header_rows, body_rows = [], []
            in_body = False
            prev_data_types = None
            for row in sheet.iter_rows():
                datarow = [ str(cell.value) for cell in row ]
                curr_data_types = [ type(cell.value) for cell in row ]
                if not in_body and prev_data_types and curr_data_types != prev_data_types:
                    in_body = True
                if in_body:
                    body_rows.append(datarow)
                else:
                    header_rows.append(datarow)
                prev_data_types = curr_data_types
            if body_rows:
                formatter = GridFormatterWithHeader(header_rows, body_rows, sheet.max_column)
            else:
                formatter = GridFormatter(header_rows, sheet.max_column)
            print(formatter)